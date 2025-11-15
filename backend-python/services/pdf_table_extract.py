"""
PDF Table Extraction Service
Handles table detection and extraction from PDFs using Camelot
"""

import logging
from pathlib import Path
from typing import Dict, Any, List, Optional
import camelot
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd

from config import Settings
from exceptions import PDFProcessingError, ValidationError

logger = logging.getLogger(__name__)


class PdfTableExtractService:
    """Service for extracting tables from PDFs using Camelot"""

    def __init__(self, settings: Settings):
        self.settings = settings
        self.logger = logger
        self.logger.info("PdfTableExtractService initialized successfully")

    def _split_merged_table(self, df: pd.DataFrame) -> List[pd.DataFrame]:
        """
        Split a single merged table into multiple logical tables.

        Detects table boundaries by identifying rows that start new logical sections,
        such as "Reference|Product" (products table) or "Tax Detail" (tax summary table).

        Args:
            df: DataFrame that may contain multiple merged tables

        Returns:
            List of DataFrames, one for each detected table
        """
        if len(df) < 3:
            return [df]

        split_indices = []

        for i in range(1, len(df)):  # Check from row 1 onward
            row = df.iloc[i]

            # Get the text content of the first two columns (most discriminative)
            col0_text = str(row.iloc[0] if not pd.isna(row.iloc[0]) else '').lower().strip()
            col1_text = str(row.iloc[1] if len(row) > 1 and not pd.isna(row.iloc[1]) else '').lower().strip()

            # Pattern 1: "Reference | Product..." indicates start of products table
            if col0_text == 'reference' and 'product' in col1_text:
                if not split_indices or split_indices[-1] < i - 1:
                    split_indices.append(i)
                    self.logger.debug(f"Found table boundary at row {i}: Products table header")

            # Pattern 2: "Tax Detail" indicates start of tax summary table
            elif 'tax' in col0_text and 'detail' in col0_text:
                if not split_indices or split_indices[-1] < i - 1:
                    split_indices.append(i)
                    self.logger.debug(f"Found table boundary at row {i}: Tax Detail table header")

            # Pattern 3: "Payment Method" indicates start of payment info section
            elif 'payment' in col0_text and 'method' in col0_text:
                if not split_indices or split_indices[-1] < i - 1:
                    split_indices.append(i)
                    self.logger.debug(f"Found table boundary at row {i}: Payment table header")

        # If no split points found, return original table
        if not split_indices:
            self.logger.debug("No table boundaries detected - returning as single table")
            return [df]

        # Split the dataframe at detected boundaries
        tables = []
        start_idx = 0

        for split_idx in split_indices:
            # Extract the table segment
            table_segment = df.iloc[start_idx:split_idx].copy()

            # Remove trailing rows that are mostly empty (> 70% empty cells)
            while len(table_segment) > 0:
                last_row = table_segment.iloc[-1]
                empty_count = sum(1 for val in last_row if pd.isna(val) or str(val).strip() == '')
                if empty_count / len(last_row) > 0.7:
                    table_segment = table_segment.iloc[:-1]
                else:
                    break

            # Only add if table has content
            if len(table_segment) > 0:
                table_segment = table_segment.reset_index(drop=True)
                tables.append(table_segment)
                self.logger.debug(f"Created table segment: {len(table_segment)} rows")

            start_idx = split_idx

        # Don't forget the last segment
        last_segment = df.iloc[start_idx:].copy()

        # Remove leading empty rows
        while len(last_segment) > 0:
            first_row = last_segment.iloc[0]
            empty_count = sum(1 for val in first_row if pd.isna(val) or str(val).strip() == '')
            if empty_count / len(first_row) > 0.7:
                last_segment = last_segment.iloc[1:]
            else:
                break

        if len(last_segment) > 0:
            last_segment = last_segment.reset_index(drop=True)
            tables.append(last_segment)
            self.logger.debug(f"Created final table segment: {len(last_segment)} rows")

        self.logger.info(f"Split merged table into {len(tables)} logical tables")
        return tables if tables else [df]

    def _tables_overlap(self, table1, table2, threshold: float = 0.5) -> bool:
        """
        Check if two tables overlap significantly on the same page

        Args:
            table1: First Camelot table
            table2: Second Camelot table
            threshold: Overlap threshold (0.5 = 50% overlap)

        Returns:
            True if tables overlap significantly
        """
        try:
            # Get table dimensions (Camelot stores bbox as (x1, y1, x2, y2))
            if not hasattr(table1, '_bbox') or not hasattr(table2, '_bbox'):
                # If bbox not available, check row/column overlap
                rows_overlap = abs(len(table1.df) - len(table2.df)) < 3
                cols_overlap = abs(len(table1.df.columns) - len(table2.df.columns)) < 2
                return rows_overlap and cols_overlap

            x1_min, y1_min, x1_max, y1_max = table1._bbox
            x2_min, y2_min, x2_max, y2_max = table2._bbox

            # Calculate intersection
            x_overlap = max(0, min(x1_max, x2_max) - max(x1_min, x2_min))
            y_overlap = max(0, min(y1_max, y2_max) - max(y1_min, y2_min))

            if x_overlap == 0 or y_overlap == 0:
                return False

            intersection_area = x_overlap * y_overlap
            area1 = (x1_max - x1_min) * (y1_max - y1_min)
            area2 = (x2_max - x2_min) * (y2_max - y2_min)

            # Calculate overlap ratio
            overlap_ratio = intersection_area / min(area1, area2)

            return overlap_ratio > threshold
        except Exception as e:
            self.logger.warning(f"Error checking table overlap: {str(e)}")
            # Fallback: compare dimensions
            rows_similar = abs(len(table1.df) - len(table2.df)) < 3
            cols_similar = abs(len(table1.df.columns) - len(table2.df.columns)) < 2
            return rows_similar and cols_similar

    async def extract_tables(
        self,
        input_path: Path,
        output_dir: Path,
        output_filename: Optional[str] = None,
        include_headers: bool = True,
        pages: str = 'all'
    ) -> Dict[str, Any]:
        """
        Extract tables from a PDF file and save to Excel
        Automatically detects and uses both lattice and stream methods to find all tables.

        Args:
            input_path: Path to input PDF file
            output_dir: Directory to save output Excel file
            output_filename: Optional custom output filename (without extension)
            include_headers: Whether to include table headers in output
            pages: Pages to extract from (e.g., '1', '1,3,5', '1-end', 'all')

        Returns:
            Dictionary with extraction results
        """
        try:
            self.logger.info(f"Starting automatic table extraction from: {input_path}")
            self.logger.info(f"Pages: {pages}, Include headers: {include_headers}")

            # Step 1: Try both detection methods and combine results
            self.logger.info("Detecting tables using automatic method detection...")

            all_tables = []

            # Try lattice method first (for bordered tables)
            try:
                self.logger.info("Trying lattice method for bordered tables...")
                lattice_tables = camelot.read_pdf(
                    str(input_path),
                    pages=pages,
                    flavor='lattice',
                    suppress_stdout=True
                )
                if len(lattice_tables) > 0:
                    self.logger.info(f"Lattice method found {len(lattice_tables)} table(s)")
                    for table in lattice_tables:
                        all_tables.append({
                            'table': table,
                            'method': 'lattice',
                            'page': table.page
                        })
            except Exception as e:
                self.logger.warning(f"Lattice method failed: {str(e)}")

            # Try stream method (for non-bordered tables)
            try:
                self.logger.info("Trying stream method for non-bordered tables...")
                stream_tables = camelot.read_pdf(
                    str(input_path),
                    pages=pages,
                    flavor='stream',
                    suppress_stdout=True
                )
                if len(stream_tables) > 0:
                    self.logger.info(f"Stream method found {len(stream_tables)} table(s)")
                    for table in stream_tables:
                        # Check if this table overlaps with any lattice table
                        is_duplicate = False
                        for existing in all_tables:
                            if (existing['page'] == table.page and
                                self._tables_overlap(existing['table'], table)):
                                # Keep the one with higher accuracy
                                if table.accuracy > existing['table'].accuracy:
                                    self.logger.info(f"Replacing table on page {table.page} with higher accuracy stream result")
                                    existing['table'] = table
                                    existing['method'] = 'stream'
                                is_duplicate = True
                                break

                        if not is_duplicate:
                            all_tables.append({
                                'table': table,
                                'method': 'stream',
                                'page': table.page
                            })
            except Exception as e:
                self.logger.warning(f"Stream method failed: {str(e)}")

            # Sort tables by page number
            all_tables.sort(key=lambda x: x['page'])

            self.logger.info(f"Initial detection found {len(all_tables)} table(s)")

            # Step 1.5: Split merged tables into logical sub-tables
            expanded_tables = []
            for table_info in all_tables:
                table = table_info['table']
                df = table.df

                # Try to split this table into multiple logical tables
                split_dfs = self._split_merged_table(df)

                if len(split_dfs) > 1:
                    self.logger.info(f"Split table on page {table.page} into {len(split_dfs)} logical tables")

                for i, split_df in enumerate(split_dfs):
                    # Create a new table entry for each split
                    expanded_tables.append({
                        'table': table,  # Keep reference to original table
                        'df': split_df,   # Use the split DataFrame
                        'method': table_info['method'],
                        'page': table.page,
                        'split_index': i
                    })

            # Replace all_tables with expanded tables
            all_tables = expanded_tables

            table_count = len(all_tables)
            self.logger.info(f"Total detected {table_count} table(s) after smart splitting")

            if table_count == 0:
                raise ValidationError("No tables found in the PDF. Please ensure the PDF contains tables.")

            # Step 2: Extract tables with headers to Excel (all in one sheet)
            self.logger.info("Extracting tables to Excel (single sheet)...")

            # Generate output filename
            if output_filename:
                excel_filename = f"{output_filename}.xlsx"
            else:
                base_name = input_path.stem
                excel_filename = f"{base_name}_tables.xlsx"

            output_path = output_dir / excel_filename

            # Create Excel workbook with single sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "All Tables"

            total_rows = 0
            table_details = []
            current_row = 1  # Track current row position in sheet

            for idx, table_info in enumerate(all_tables, start=1):
                # Get table as DataFrame
                table = table_info['table']
                method = table_info['method']
                # Use split DataFrame if available, otherwise use original
                df = table_info.get('df', table.df)

                # Add table number and metadata as a header
                metadata_cell = ws.cell(row=current_row, column=1,
                                       value=f"Table {idx} (Page {table.page}, Method: {method}, Accuracy: {table.accuracy:.1f}%)")
                metadata_cell.font = Font(bold=True, size=12, color="1F4E78")
                current_row += 1

                # Determine if first row is header
                if include_headers and len(df) > 0:
                    # Use first row as header
                    headers = df.iloc[0].tolist()
                    data_df = df.iloc[1:].reset_index(drop=True)
                    data_df.columns = headers
                else:
                    data_df = df
                    headers = [f"Column_{i+1}" for i in range(len(df.columns))]

                # Write headers with styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")

                for col_idx, header in enumerate(headers, start=1):
                    cell = ws.cell(row=current_row, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                current_row += 1

                # Write data rows
                if include_headers and len(df) > 0:
                    for row_idx, row in data_df.iterrows():
                        for col_idx, value in enumerate(row, start=1):
                            ws.cell(row=current_row, column=col_idx, value=value)
                        current_row += 1
                    total_rows += len(data_df)
                else:
                    for row_idx, row in df.iterrows():
                        for col_idx, value in enumerate(row, start=1):
                            ws.cell(row=current_row, column=col_idx, value=value)
                        current_row += 1
                    total_rows += len(df)

                # Store table metadata
                table_details.append({
                    "table_number": idx,
                    "page": table.page,
                    "rows": len(df),
                    "columns": len(df.columns),
                    "accuracy": round(table.accuracy, 2),
                    "method": method
                })

                self.logger.info(
                    f"Table {idx}: Page {table.page}, "
                    f"{len(df)} rows × {len(df.columns)} cols, "
                    f"Accuracy: {table.accuracy:.2f}%, "
                    f"Method: {method}"
                )

                # Add 3 empty rows as separator (unless it's the last table)
                if idx < len(all_tables):
                    current_row += 3

            # Auto-adjust column widths for the entire sheet
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save Excel file
            wb.save(output_path)
            file_size = output_path.stat().st_size

            self.logger.info(f"Table extraction completed: {excel_filename}")
            self.logger.info(f"Total tables: {table_count}, Total rows: {total_rows}")

            return {
                "filename": excel_filename,
                "size": file_size,
                "tables_count": table_count,
                "total_rows": total_rows,
                "table_details": table_details,
                "pages": pages
            }

        except Exception as e:
            self.logger.error(f"Table extraction failed: {str(e)}")
            raise PDFProcessingError("Table Extraction", str(e))

        finally:
            # Clean up uploaded file
            if input_path.exists():
                input_path.unlink()
                self.logger.info(f"Cleaned up temporary file: {input_path}")
